# -*- coding: utf-8 -*-
# Copyright (c) 2021, libracore and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
import six
import json
from six import BytesIO
import openpyxl
from openpyxl.styles import Font

@frappe.whitelist()
def get_data(suchparameter, exportieren=False):
    if isinstance(suchparameter, six.string_types):
        suchparameter = json.loads(suchparameter)
    
    if not exportieren:
        # show data
        data = get_datas(suchparameter)
        if data:
            if suchparameter["ansicht_auswahl"] == 'SalesHeader':
                return frappe.render_template('templates/navision_export/navision_exp_salesheader.html', {'invoices': data})
            else:
                return frappe.render_template('templates/navision_export/navision_exp_salesline.html', {'invoices': data})
        else:
            return False
    else:
        if exist_navisionexport_folder():
            # SalesHeader
            suchparameter["ansicht_auswahl"] = 'SalesHeader'
            salesheader_raw_data = get_datas(suchparameter)
            salesheader_data = [["Belegart","Nr.","Verk. an Deb.-Nr.","Rech. an Deb.-Nr.","Buchungsdatum","Buchungsbeschreibung","Fälligkeitsdatum","Shortcutdimensionscode 1","Belegdatum","Externe Belegnummer","Datum der Leistung von","Datum der Leistung bis"]]
            
            for sinv in salesheader_raw_data:
                data = []
                data.append("Rechnung")
                data.append("212ERP" + sinv["sinv"].replace("RE", ""))
                data.append(sinv["navision_kundennummer"])
                data.append(sinv["navision_kundennummer"])
                data.append(frappe.utils.get_datetime(sinv["rechnungsdatum"]).strftime('%d.%m.%Y'))
                data.append(sinv["buchungsbeschreibung"])
                data.append(frappe.utils.get_datetime(sinv["due_date"]).strftime('%d.%m.%Y'))
                data.append(sinv["navision_shortcutdimensionscode_1"])
                data.append(frappe.utils.get_datetime(sinv["rechnungsdatum"]).strftime('%d.%m.%Y'))
                data.append(sinv["sinv"])
                salesheader_data.append(data)
            
            # SalesLine
            suchparameter["ansicht_auswahl"] = 'SalesLine'
            salesline_raw_data = get_datas(suchparameter)
            salesline_data = [["Belegart","Belegnr.","Zeilennr.","Art","Nr.","Lagerortcode","Beschreibung","Einheit","Menge","VK-Preis","Shortcutdimensionscode 1","MwSt.-Geschäftsbuchungsgruppe","MwSt.-Produktbuchungsgruppe","Einheitencode"]]
            for _salesline_data in salesline_raw_data:
                salesline_data.append(_salesline_data)
            
            xlsx_file = make_navision_xlsx(salesheader_data, salesline_data)
            file_data = xlsx_file.getvalue()
            
            _file = frappe.get_doc({
            "doctype": "File",
            "file_name": "{title}_{date_von}_{date_bis}.xlsx".format(date_von=suchparameter["date_von"], date_bis=suchparameter["date_bis"], title='SalesHeader'),
            "folder": "Home/NavisionExport",
            "is_private": 1,
            "content": file_data})
            _file.save()
            
            return True
    # fallback
    return False

def exist_navisionexport_folder():
    exist = frappe.db.sql("""SELECT COUNT(`name`) AS `qty` FROM `tabFile` WHERE `name` = 'Home/NavisionExport' AND `is_folder` = 1""", as_dict=True)
    if exist[0].qty != 1:
        new_folder = frappe.get_doc({
            "doctype": "File",
            "file_name": "NavisionExport",
            "folder": "Home",
            "is_folder": 1
        })
        new_folder.insert()
        frappe.db.commit()
    return True
        
        
def make_navision_xlsx(salesheader_data, salesline_data):

    wb = openpyxl.Workbook(write_only=True)
    
    # SalesHeader
    ws = wb.create_sheet('SalesHeader', 0)
    row1 = ws.row_dimensions[1]
    row1.font = Font(name='Calibri',bold=False)
    for row in salesheader_data:
        clean_row = []
        for item in row:
            clean_row.append(item)

        ws.append(clean_row)
        
    # SalesLine
    ws = wb.create_sheet('SalesLine', 1)
    row1 = ws.row_dimensions[1]
    row1.font = Font(name='Calibri',bold=False)
    for row in salesline_data:
        clean_row = []
        for item in row:
            clean_row.append(item)

        ws.append(clean_row)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def get_datas(suchparameter):
    if suchparameter["ansicht_auswahl"] == 'SalesHeader':
        return _get_salesheader_datas(suchparameter)
    else:
        return _get_salesline_datas(suchparameter)

def _get_salesheader_datas(suchparameter):
    # SalesHeader
    sinvs = frappe.db.sql("""SELECT `name`, `customer`, `posting_date`, `due_date`, `billing_type`, `project`, `cost_center` FROM `tabSales Invoice` WHERE `posting_date` BETWEEN '{date_von}' AND '{date_bis}' AND `docstatus` = 1 AND `is_return` != 1 AND `rechnung_nach_navision_exportiert` != 1""".format(date_von=suchparameter["date_von"], date_bis=suchparameter["date_bis"]), as_dict=True)
    datas = []
    for sinv in sinvs:
        cost_center = frappe.get_doc("Cost Center", sinv.cost_center)
        
        if sinv.project:
            projekt = frappe.get_doc("Project", sinv.project)
            projektnummer_rhapsody = projekt.projektnummer_rhapsody
            contract_type = projekt.contract_type
            if contract_type == 'Werkvertrag':
                contract_type = 'WV'
            else:
                contract_type = 'DV'
            projektnummer = str(int(str(projekt.name).replace("P", "")))
        else:
            projekt = False
            projektnummer_rhapsody = False
            contract_type = 'DV'
            projektnummer = False
        
        navision_kundennummer = frappe.get_doc("Customer", sinv.customer).navision_nr
        
        buchungsbeschreibung = ''
        if sinv.billing_type == 'Rechnung':
            buchungsbeschreibung += sinv.billing_type + " " + contract_type
        else:
            if sinv.billing_type == 'Teilrechnung':
                xte_rechnung = frappe.db.sql("""SELECT `idx` FROM `tabSales Order Anzahlung` WHERE `sales_invoice` = '{sinv}'""".format(sinv=sinv.name), as_dict=True)
                if len(xte_rechnung) > 0:
                    buchungsbeschreibung += '{xte_rechnung}.'.format(xte_rechnung=xte_rechnung[0].idx)
                buchungsbeschreibung += " TR"
            if sinv.billing_type == 'Schlussrechnung':
                buchungsbeschreibung += "SR"
            if projekt:
                buchungsbeschreibung += " zu P{projektnummer}".format(projektnummer=projektnummer)
                if projektnummer_rhapsody:
                    buchungsbeschreibung += " RhNr. {projektnummer_rhapsody}".format(projektnummer_rhapsody=projektnummer_rhapsody)
        
        data = {
            'sinv': sinv.name,
            'navision_kundennummer': navision_kundennummer,
            'rechnungsdatum': sinv.posting_date,
            'buchungsbeschreibung': buchungsbeschreibung,
            'due_date': sinv.due_date,
            'navision_shortcutdimensionscode_1': cost_center.navision_shortcutdimensionscode_1 if cost_center.navision_shortcutdimensionscode_1 else '1000800'
        }
        
        datas.append(data)
        
    if len(datas) > 0:
        return datas
    else:
        return False

def _get_salesline_datas(suchparameter):
    # SalesLine
    invoices = _get_salesheader_datas(suchparameter)
    
    datas = []
    for _sinv in invoices:
        sinv = frappe.get_doc("Sales Invoice", _sinv["sinv"])
        if sinv.billing_type == 'Rechnung':
            loop = 0
            for lineitem in sinv.items:
                data = []
                data.append("Rechnung")
                data.append("212ERP" + sinv.name.replace("RE", ""))
                data.append("1000" + str(loop))
                loop += 1
                data.append("Sachkonto")
                data.append(sinv.navision_kontonummer)
                data.append("")
                data.append(lineitem.item_name[:50])
                data.append("")
                data.append(lineitem.qty)
                data.append(lineitem.rate)
                data.append(_sinv["navision_shortcutdimensionscode_1"])
                data.append("INL")
                if len(sinv.taxes) > 0:
                    data.append(sinv.taxes[0].rate)
                else:
                    data.append("")
                data.append("")
                datas.append(data)
        if sinv.billing_type == 'Teilrechnung':
            xte_rechnung = frappe.db.sql("""SELECT `idx` FROM `tabSales Order Anzahlung` WHERE `sales_invoice` = '{sinv}'""".format(sinv=sinv.name), as_dict=True)
            if len(xte_rechnung) > 0:
                xte_rechnung = str(xte_rechnung[0].idx) + ". "
            else:
                xte_rechnung = ''
            data = []
            data.append("Rechnung")
            data.append("212ERP" + sinv.name.replace("RE", ""))
            data.append("10000")
            data.append("Sachkonto")
            data.append("17200")
            data.append("")
            data.append(_sinv["buchungsbeschreibung"])
            data.append("")
            data.append("1")
            data.append(sinv.grand_total)
            data.append(_sinv["navision_shortcutdimensionscode_1"])
            data.append("INL")
            data.append("0")
            data.append("")
            datas.append(data)
        if sinv.billing_type == 'Schlussrechnung':
            loop = 0
            for lineitem in sinv.items:
                data = []
                data.append("Rechnung")
                data.append("212ERP" + sinv.name.replace("RE", ""))
                data.append("1000" + str(loop))
                loop += 1
                data.append("Sachkonto")
                data.append(sinv.navision_kontonummer)
                data.append("")
                data.append(lineitem.item_name[:50])
                data.append("")
                data.append(lineitem.qty)
                data.append(lineitem.rate)
                data.append(_sinv["navision_shortcutdimensionscode_1"])
                data.append("INL")
                if len(sinv.taxes) > 0:
                    data.append(sinv.taxes[0].rate)
                else:
                    data.append("")
                data.append("")
                datas.append(data)
            sales_order = frappe.get_doc("Sales Order", sinv.items[0].sales_order)
            for teilrechnung in sales_order.billing_overview:
                if teilrechnung.sales_invoice != sinv.name:
                    _teilrechnung = frappe.get_doc("Sales Invoice", teilrechnung.sales_invoice)
                    data = []
                    data.append("Rechnung")
                    data.append("212ERP" + sinv.name.replace("RE", ""))
                    data.append("1000" + str(loop))
                    loop += 1
                    data.append("Sachkonto")
                    data.append("17100")
                    data.append("")
                    if not teilrechnung.invoice_rhapsody:
                        data.append("Abschlag RE Nr. " + teilrechnung.sales_invoice + " vom " + str(_teilrechnung.posting_date))
                    else:
                        data.append("Abschlag RE Nr. " + teilrechnung.invoice_rhapsody + " vom " + str(_teilrechnung.posting_date))
                    data.append("")
                    data.append("1")
                    data.append("-" + str(teilrechnung.amount))
                    data.append(_sinv["navision_shortcutdimensionscode_1"])
                    data.append("INL")
                    if len(sinv.taxes) > 0:
                        data.append(sinv.taxes[0].rate)
                    else:
                        data.append("")
                    data.append("")
                    datas.append(data)
    if len(datas) > 0:
        return datas
    else:
        return False
